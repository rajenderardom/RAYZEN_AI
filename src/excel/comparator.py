"""
RAYZEN AI
Excel Comparator

Version : 0.1.0
"""

import os
from typing import List, Optional, Dict, Any
import openpyxl
from openpyxl.styles import PatternFill
from src.excel.manager import ExcelManager
from src.core.logger import RayzenLogger


class ExcelComparator:
    """Provides capabilities to compare two Excel workbooks and identify differences."""

    def __init__(self, excel_manager_factory=None):
        """Initialize ExcelComparator.

        Args:
            excel_manager_factory: Optional callable returning ExcelManager instances.
        """
        self.logger = RayzenLogger()
        self.excel_manager_factory = excel_manager_factory or ExcelManager

    def compare(
        self,
        file_path_old: str,
        file_path_new: str,
        key_column: str,
        sheet_name: Optional[str] = None,
        export_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Compare two Excel files based on a unique key column.

        Args:
            file_path_old (str): Path to the old Excel file.
            file_path_new (str): Path to the new Excel file.
            key_column (str): Unique column identifier to match rows.
            sheet_name (Optional[str]): Sheet to compare. Defaults to active sheet.
            export_path (Optional[str]): Path to save the comparison report.

        Returns:
            Dict[str, Any]: Comparison summary statistics.
        """
        self.logger.info(
            f"Starting Excel comparison: '{file_path_old}' vs '{file_path_new}' using key: '{key_column}'"
        )

        if not os.path.exists(file_path_old):
            self.logger.error(f"Old file does not exist: {file_path_old}")
            return {"success": False, "error": f"Old file not found: {file_path_old}"}

        if not os.path.exists(file_path_new):
            self.logger.error(f"New file does not exist: {file_path_new}")
            return {"success": False, "error": f"New file not found: {file_path_new}"}

        manager_old = self.excel_manager_factory()
        manager_new = self.excel_manager_factory()

        try:
            if not manager_old.open_workbook(file_path_old):
                raise RuntimeError(f"Failed to open old workbook: {file_path_old}")
            if not manager_new.open_workbook(file_path_new):
                raise RuntimeError(f"Failed to open new workbook: {file_path_new}")

            # Resolve sheet name
            sh_old_name = sheet_name or manager_old.get_active_sheet_name()
            sh_new_name = sheet_name or manager_new.get_active_sheet_name()

            if not sh_old_name or sh_old_name not in manager_old.get_sheet_names():
                raise ValueError(f"Sheet '{sh_old_name}' not found in old workbook.")
            if not sh_new_name or sh_new_name not in manager_new.get_sheet_names():
                raise ValueError(f"Sheet '{sh_new_name}' not found in new workbook.")

            sheet_old = manager_old.wb[sh_old_name]
            sheet_new = manager_new.wb[sh_new_name]

            # Read headers
            old_headers = [cell.value for cell in sheet_old[1] if cell.value is not None]
            new_headers = [cell.value for cell in sheet_new[1] if cell.value is not None]

            if key_column not in old_headers:
                raise ValueError(f"Key column '{key_column}' not found in old workbook headers.")
            if key_column not in new_headers:
                raise ValueError(f"Key column '{key_column}' not found in new workbook headers.")

            key_idx_old = old_headers.index(key_column)
            key_idx_new = new_headers.index(key_column)

            # Load rows
            old_data: Dict[Any, Dict[str, Any]] = {}
            for row in sheet_old.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= key_idx_old:
                    continue
                key_val = row[key_idx_old]
                if key_val is None:
                    continue
                row_dict = {
                    old_headers[i]: row[i] for i in range(min(len(row), len(old_headers)))
                }
                old_data[key_val] = row_dict

            new_data: Dict[Any, Dict[str, Any]] = {}
            for row in sheet_new.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= key_idx_new:
                    continue
                key_val = row[key_idx_new]
                if key_val is None:
                    continue
                row_dict = {
                    new_headers[i]: row[i] for i in range(min(len(row), len(new_headers)))
                }
                new_data[key_val] = row_dict

            # Analyze changes
            added_keys = []
            removed_keys = []
            modified_keys = []
            modifications: Dict[Any, Dict[str, Dict[str, Any]]] = {}

            # Added records
            for key in new_data:
                if key not in old_data:
                    added_keys.append(key)

            # Removed records
            for key in old_data:
                if key not in new_data:
                    removed_keys.append(key)

            # Modified records
            shared_cols = [col for col in new_headers if col in old_headers]
            for key in new_data:
                if key in old_data:
                    row_old = old_data[key]
                    row_new = new_data[key]
                    changes = {}
                    for col in shared_cols:
                        if row_old.get(col) != row_new.get(col):
                            changes[col] = {
                                "old": row_old.get(col),
                                "new": row_new.get(col),
                            }
                    if changes:
                        modified_keys.append(key)
                        modifications[key] = changes

            summary = {
                "success": True,
                "total_old": len(old_data),
                "total_new": len(new_data),
                "added_count": len(added_keys),
                "removed_count": len(removed_keys),
                "modified_count": len(modified_keys),
                "added_records": added_keys,
                "removed_records": removed_keys,
                "modified_records": modified_keys,
            }

            self.logger.info(
                f"Comparison statistics: Added={len(added_keys)}, Removed={len(removed_keys)}, Modified={len(modified_keys)}"
            )

            # Export comparison report if requested
            if export_path:
                self._export_report(
                    export_path=export_path,
                    key_column=key_column,
                    new_headers=new_headers,
                    old_headers=old_headers,
                    new_data=new_data,
                    old_data=old_data,
                    added_keys=added_keys,
                    removed_keys=removed_keys,
                    modified_keys=modified_keys,
                    modifications=modifications,
                    summary=summary,
                )

            return summary

        except Exception as e:
            self.logger.error(f"Error occurred during Excel comparison: {e}")
            return {"success": False, "error": str(e)}
        finally:
            manager_old.close_workbook()
            manager_new.close_workbook()

    def _export_report(
        self,
        export_path: str,
        key_column: str,
        new_headers: List[str],
        old_headers: List[str],
        new_data: Dict[Any, Dict[str, Any]],
        old_data: Dict[Any, Dict[str, Any]],
        added_keys: List[Any],
        removed_keys: List[Any],
        modified_keys: List[Any],
        modifications: Dict[Any, Dict[str, Dict[str, Any]]],
        summary: Dict[str, Any],
    ) -> None:
        """Helper to create and save comparison workbook report with color highlighting."""
        self.logger.info(f"Exporting comparison report to: {export_path}")
        os.makedirs(os.path.dirname(os.path.abspath(export_path)), exist_ok=True)

        wb_report = openpyxl.Workbook()

        # 1. Summary Sheet
        ws_summary = wb_report.active
        ws_summary.title = "Summary"
        ws_summary.append(["RAYZEN AI - Excel Comparison Summary"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Old Records", summary["total_old"]])
        ws_summary.append(["Total New Records", summary["total_new"]])
        ws_summary.append(["Added Records", summary["added_count"]])
        ws_summary.append(["Removed Records", summary["removed_count"]])
        ws_summary.append(["Modified Records", summary["modified_count"]])

        # Style summary header
        ws_summary["A1"].font = openpyxl.styles.Font(size=14, bold=True)
        ws_summary["A3"].font = openpyxl.styles.Font(bold=True)
        ws_summary["B3"].font = openpyxl.styles.Font(bold=True)

        # 2. Details Sheet
        ws_details = wb_report.create_sheet(title="Comparison Details")

        headers = list(new_headers)
        for col in old_headers:
            if col not in headers:
                headers.append(col)
        headers.append("Comparison Status")
        ws_details.append(headers)

        # Styles
        fill_added = PatternFill(
            start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
        )  # Light green
        fill_removed = PatternFill(
            start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
        )  # Light red
        fill_modified = PatternFill(
            start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
        )  # Light yellow

        # Add records from new sheet (includes added, modified, unchanged)
        for key, row_dict in new_data.items():
            row_values = []
            for col in headers[:-1]:
                row_values.append(row_dict.get(col))

            status = "Unchanged"
            if key in added_keys:
                status = "Added"
            elif key in modified_keys:
                status = "Modified"

            row_values.append(status)
            ws_details.append(row_values)
            curr_row = ws_details.max_row

            # Apply highlights
            if status == "Added":
                for col_idx in range(1, len(headers) + 1):
                    ws_details.cell(row=curr_row, column=col_idx).fill = fill_added
            elif status == "Modified":
                ws_details.cell(row=curr_row, column=len(headers)).fill = fill_modified
                changes = modifications.get(key, {})
                for col_name in changes:
                    if col_name in headers:
                        c_idx = headers.index(col_name) + 1
                        ws_details.cell(row=curr_row, column=c_idx).fill = fill_modified

        # Add removed records at the end
        for key in removed_keys:
            row_dict = old_data[key]
            row_values = []
            for col in headers[:-1]:
                row_values.append(row_dict.get(col))
            row_values.append("Removed")
            ws_details.append(row_values)
            curr_row = ws_details.max_row

            for col_idx in range(1, len(headers) + 1):
                ws_details.cell(row=curr_row, column=col_idx).fill = fill_removed

        wb_report.save(export_path)
        self.logger.info("Comparison report exported successfully.")
