import os
import shutil
import tempfile
import unittest
import openpyxl
from src.excel.comparator import ExcelComparator
from src.core.skill_registry import SkillRegistry


class TestExcelComparator(unittest.TestCase):
    """Test cases for ExcelComparator."""

    def setUp(self):
        self.test_dir = tempfile.mkdtemp()
        self.old_path = os.path.join(self.test_dir, "old.xlsx")
        self.new_path = os.path.join(self.test_dir, "new.xlsx")
        self.export_path = os.path.join(self.test_dir, "report.xlsx")
        self.comparator = ExcelComparator()

        # Build Old Workbook
        wb_old = openpyxl.Workbook()
        ws_old = wb_old.active
        ws_old.title = "Sheet1"
        ws_old.append(["Site ID", "Name", "Status"])
        ws_old.append(["S001", "Old Name 1", "Active"])
        ws_old.append(["S002", "Name 2", "Inactive"])
        ws_old.append(["S003", "Name 3", "Active"])
        wb_old.save(self.old_path)

        # Build New Workbook
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = "Sheet1"
        ws_new.append(["Site ID", "Name", "Status"])
        ws_new.append(["S001", "New Name 1", "Active"])  # Modified Name
        ws_new.append(["S002", "Name 2", "Inactive"])  # Unchanged
        ws_new.append(["S004", "Name 4", "Active"])  # Added S004 (S003 Removed)
        wb_new.save(self.new_path)

    def tearDown(self):
        shutil.rmtree(self.test_dir)

    def test_compare_success(self):
        result = self.comparator.compare(
            file_path_old=self.old_path,
            file_path_new=self.new_path,
            key_column="Site ID",
            sheet_name="Sheet1",
            export_path=self.export_path,
        )

        self.assertTrue(result["success"])
        self.assertEqual(result["total_old"], 3)
        self.assertEqual(result["total_new"], 3)
        self.assertEqual(result["added_count"], 1)
        self.assertEqual(result["removed_count"], 1)
        self.assertEqual(result["modified_count"], 1)

        self.assertEqual(result["added_records"], ["S004"])
        self.assertEqual(result["removed_records"], ["S003"])
        self.assertEqual(result["modified_records"], ["S001"])

        # Check export report exists
        self.assertTrue(os.path.exists(self.export_path))

        # Verify contents of exported report
        wb_rep = openpyxl.load_workbook(self.export_path)
        self.assertIn("Summary", wb_rep.sheetnames)
        self.assertIn("Comparison Details", wb_rep.sheetnames)

        ws_det = wb_rep["Comparison Details"]
        rows = list(ws_det.iter_rows(values_only=True))
        self.assertEqual(rows[0], ("Site ID", "Name", "Status", "Comparison Status"))

        # S001 should be Modified
        row1 = rows[1]
        self.assertEqual(row1[0], "S001")
        self.assertEqual(row1[3], "Modified")

        # Check cell background fill of modified Name cell
        cell_name = ws_det.cell(row=2, column=2)
        self.assertEqual(cell_name.fill.start_color.rgb, "00FFF3CD")

    def test_compare_key_column_missing(self):
        result = self.comparator.compare(
            file_path_old=self.old_path,
            file_path_new=self.new_path,
            key_column="Nonexistent Column",
            sheet_name="Sheet1",
        )
        self.assertFalse(result["success"])
        self.assertIn("Key column", result["error"])

    def test_compare_missing_file(self):
        result = self.comparator.compare(
            file_path_old="invalid_path.xlsx", file_path_new=self.new_path, key_column="Site ID"
        )
        self.assertFalse(result["success"])
        self.assertIn("Old file not found", result["error"])

    def test_skill_registry_integration(self):
        registry = SkillRegistry()
        registry.register("excel_comparator", self.comparator)

        result = registry.execute(
            "excel_comparator",
            "compare",
            file_path_old=self.old_path,
            file_path_new=self.new_path,
            key_column="Site ID",
            sheet_name="Sheet1",
        )
        self.assertTrue(result["success"])
        self.assertEqual(result["added_count"], 1)


if __name__ == "__main__":
    unittest.main()
